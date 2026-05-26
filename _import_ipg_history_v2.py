#!/usr/bin/env python3
"""V2 importer — reads ALL sheets in IPG Bet history.xlsx (0.5G, 1.5G, 2.5G).
Outputs records keyed by sheet's market tag, ready to merge."""
import openpyxl, json, sys
from datetime import datetime, timezone

XLSX = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\toddb\Downloads\IPG Bet history.xlsx'
OUT  = sys.argv[2] if len(sys.argv) > 2 else r'C:\Users\toddb\OneDrive\Desktop\telegram\ipg_history_records_v2.json'

wb = openpyxl.load_workbook(XLSX)
records = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, system, market, match, odds, wl, pl = row
        if not system or not match:
            continue
        sys_clean = str(system).strip()
        if not sys_clean.lower().startswith('system'):
            sys_clean = 'System' + sys_clean
        market_str = str(market) if market is not None else sheet_name.replace('G','')
        odds_f = float(odds) if odds is not None else None
        pl_f   = float(pl) if pl is not None else 0
        wl_norm = (str(wl).strip().lower())
        result = 'won' if wl_norm == 'w' else 'lost' if wl_norm == 'l' else 'void' if wl_norm == 'v' else 'pending'
        placed_at = (date if isinstance(date, datetime) else datetime.now()).replace(tzinfo=timezone.utc).isoformat().replace('+00:00','Z')
        teams = [t.strip() for t in str(match).split(' V ', 1)] if ' V ' in str(match) else [str(match), '']
        teamA = teams[0] if teams else ''
        teamB = teams[1] if len(teams) > 1 else ''
        rec_id = f"IPG-HIST-{sheet_name}-{sys_clean}-{placed_at}-{teamA[:8]}".replace(' ','_').replace(':','').replace('-','_')
        records.append({
            'id': rec_id,
            'historical': True,
            'historicalSource': 'IPG Bet history.xlsx',
            'historicalSheet': sheet_name,
            'placedAt': placed_at,
            'match': str(match),
            'teamA': teamA,
            'teamB': teamB,
            'marketName': f'Over/Under {market_str} Goals',
            'overUnderValue': market_str,
            'strategy': sys_clean,
            'signalStrategy': sys_clean,
            'overOdds': odds_f,
            'underOdds': None,
            'success': True,
            'isCompanion': False,
            'result': result,
            'avePoints': pl_f,
            'finalGoalsA': None,
            'finalGoalsB': None,
            'finalGoals': None,
            'priceSnapshots': [],
        })

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(records, f, indent=2, ensure_ascii=False)
print(f'wrote {len(records)} records from {len(wb.sheetnames)} sheets to {OUT}')

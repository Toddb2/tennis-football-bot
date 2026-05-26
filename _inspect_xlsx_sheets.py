import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\toddb\Downloads\IPG Bet history.xlsx')
print('sheets:', wb.sheetnames)
for n in wb.sheetnames:
    ws = wb[n]
    print(f'\n-- {n} -- rows={ws.max_row} cols={ws.max_column}')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
        print(' ', i, list(row))
